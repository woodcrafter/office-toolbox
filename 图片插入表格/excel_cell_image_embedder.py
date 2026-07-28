#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 图片真嵌入单元格工具
功能：根据 Excel 指定列的单元格内容匹配图片文件名，将图片真正嵌入单元格
     （Excel「放置在单元格中」格式，非悬浮图片）。
     嵌入后图片随单元格移动/筛选/排序，其他 sheet 可用 =Sheet1!B2 引用该图片。

显示要求：Microsoft 365（Excel 2021 之前的旧版 Excel 不支持），或较新版本 WPS。
"""

import os
import re
import sys
import shutil
import secrets
import zipfile
import tempfile
import datetime
import posixpath
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False


# xlsxwriter 直接支持的图片格式
SUPPORTED_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}
# 需要 Pillow 转换后才能嵌入的格式
CONVERT_EXTS = {'.webp'}

# Excel 行高上限 409.5 磅 ≈ 546 像素
MAX_ROW_HEIGHT_PT = 409.5
PX_TO_PT = 0.75  # 96dpi 像素 -> 磅


# ============== 阈值说明文本 ==============

THRESHOLD_HELP_TEXT = """【阈值说明】

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【模糊匹配处理流程】

当选择"模糊匹配"时，系统会按以下两步处理：

  第1步 — 精确匹配
    先用 Excel 单元格值直接查找图片文件名。
    如果找到，直接记录为【精确匹配成功】，不再进行模糊匹配。

  第2步 — 模糊匹配（仅第1步未匹配时执行）
    对所有图片文件名逐一评分，选择分数最高的一个。
    如果最高分 ≥ 阈值，则记录为【模糊匹配成功】。

  第3步 — 未匹配
    前两步都找不到对应图片，记录为【未匹配】。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【评分等级】

100% — 精确匹配
   Excel 单元格值与图片文件名完全一致
   （或仅空格 ↔ 下划线差异）
   误匹配风险：无

85%  — 长字符串包含匹配
   Excel 值被完整包含在图片名中，且双方长度均≥10字符
   例如：Excel="DRA343"  图片="NO.JA-QS-0403-3732_V1.0-DRA343.png"
   误匹配风险：极低

60%  — 短字符串包含匹配
   Excel 值被包含在图片名中，但某一方长度<10字符
   例如：Excel="C001"     图片="NO.JA-QS-0403-0001_V1.1-C001.png"
   误匹配风险：中等（可能出现多个图片含相同短字符串）

<60% — LCS（最长公共子串）相似度
   两个字符串有较长的公共部分，但互不为子串
   例如：Excel="ABC-123"  图片="ABC-124"
   误匹配风险：随分数降低而升高

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【推荐阈值】

90%（0.90）— 极严格
   只接受精确匹配 + 长包含匹配。误匹配极少。

60%（0.60）— 较严格（推荐）
   接受精确匹配 + 包含匹配。平衡覆盖面与准确性。

30%（0.30）— 宽松（默认）
   接受部分相似匹配。覆盖面最广，可能有少量误匹配。

10%（0.10）— 极宽松
   几乎任何有点相似的都会匹配。误匹配风险较高，
   仅建议在人工复核后使用。
"""

WPS_NOTE = (
    "说明：WPS 原生嵌入格式（DISPIMG），图片真正嵌入单元格，可用 WPS 正常查看；"
    "注意此格式在 Microsoft Excel 中无法显示图片（会显示为公式），两者不通用。"
)

EXCEL365_NOTE = (
    "说明：Excel 365「放置在单元格中」格式，需要 Microsoft 365 或支持该格式的新版 WPS 打开；"
    "旧版 Excel（2021 及更早）和部分 WPS 版本会显示 #VALUE! 或图片破损。"
)


# ============== 工具函数 ==============

def column_letter_to_index(col_letter: str) -> int:
    """将列字母（如 'A', 'AB'）转换为 1-based 列索引。"""
    col_letter = col_letter.strip().upper()
    if not col_letter:
        raise ValueError("列名不能为空")
    try:
        idx = 0
        for ch in col_letter:
            if 'A' <= ch <= 'Z':
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            else:
                raise ValueError
        return idx
    except Exception:
        raise ValueError(f"无效的列名: {col_letter}")


def get_image_files(folder: str):
    """递归获取文件夹及其所有子文件夹中支持的图片文件，
    返回 {文件名(无扩展名): 完整路径} 和 {文件名(含扩展名): 完整路径}。"""
    exts = SUPPORTED_EXTS | CONVERT_EXTS
    without_ext = {}
    with_ext = {}
    folder = os.path.abspath(folder)

    for root, _dirs, files in os.walk(folder):
        for f in files:
            p = os.path.join(root, f)
            name_lower = f.lower()
            ext = os.path.splitext(name_lower)[1]
            if ext in exts:
                if f not in with_ext:
                    with_ext[f] = p
                    without_ext[os.path.splitext(f)[0]] = p
    return with_ext, without_ext


def similarity_score(a: str, b: str) -> float:
    """最长公共子串(LCS)长度与较长字符串长度的比值 (0.0 ~ 1.0)。"""
    if not a or not b:
        return 0.0
    a, b = a.lower(), b.lower()
    m, n = len(a), len(b)
    max_len = 0
    dp = [0] * (n + 1)
    for i in range(1, m + 1):
        prev = 0
        for j in range(1, n + 1):
            temp = dp[j]
            if a[i - 1] == b[j - 1]:
                dp[j] = prev + 1
                max_len = max(max_len, dp[j])
            else:
                dp[j] = 0
            prev = temp
    longer = max(m, n)
    return max_len / longer if longer > 0 else 0.0


def find_best_fuzzy_match(key: str, candidates: dict, threshold: float = 0.3) -> tuple:
    """在候选图片中找到与 key 最相似的一个。
    返回 (img_path, score, matched_name) 或 (None, 0.0, None)。"""
    normalized_key = key.replace(' ', '_')
    best_path = None
    best_name = None
    best_score = 0.0

    for name, path in candidates.items():
        score = 0.0
        if key == name or normalized_key == name:
            score = 1.0
        elif key in name or name in key or normalized_key in name or name in normalized_key:
            if len(key) >= 10 and len(name) >= 10:
                score = 0.85
            else:
                score = 0.60
        else:
            score = similarity_score(key, name)

        if score > best_score:
            best_score = score
            best_path = path
            best_name = name

    if best_score >= threshold:
        return best_path, best_score, best_name
    return None, 0.0, None


class ImageConverter:
    """把 xlsxwriter 不支持的图片格式（如 webp）临时转换为 PNG。"""

    def __init__(self):
        self._tmpdir = None
        self._cache = {}

    def get_embeddable_path(self, img_path: str) -> str:
        ext = os.path.splitext(img_path)[1].lower()
        if ext in SUPPORTED_EXTS:
            return img_path
        if img_path in self._cache:
            return self._cache[img_path]
        if not HAS_PIL:
            raise RuntimeError(f"图片格式 {ext} 需要 Pillow 库转换后才能嵌入，请安装 Pillow。")
        if self._tmpdir is None:
            self._tmpdir = tempfile.mkdtemp(prefix="img_embed_")
        out = os.path.join(self._tmpdir, f"conv_{len(self._cache)}.png")
        with PILImage.open(img_path) as im:
            im.save(out, "PNG")
        self._cache[img_path] = out
        return out

    def get_image_size(self, img_path: str):
        """返回 (宽, 高) 像素；读取失败时返回 None。"""
        if not HAS_PIL:
            return None
        try:
            with PILImage.open(img_path) as im:
                return im.size
        except Exception:
            return None


# ============== WPS 嵌入格式（DISPIMG / cellimages.xml） ==============

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_ETC = "http://www.wps.cn/officeDocument/2017/etCustomData"
_NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"

# WPS 原生文件中确认的关键常量
WPS_CELLIMAGE_REL_TYPE = "http://www.wps.cn/officeDocument/2020/cellImage"
WPS_CELLIMAGE_CONTENT_TYPE = "application/vnd.wps-officedocument.cellimage+xml"
_IMAGE_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"

ET.register_namespace("", _NS_MAIN)
ET.register_namespace("r", _NS_R)
ET.register_namespace("xdr", _NS_XDR)
ET.register_namespace("a", _NS_A)
ET.register_namespace("etc", _NS_ETC)

_EXT_CONTENT_TYPES = {
    '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
    '.gif': 'image/gif', '.bmp': 'image/bmp', '.webp': 'image/webp',
}


def _next_rid(rel_root):
    max_id = 0
    for rel in rel_root.findall(f"{{{_NS_REL}}}Relationship"):
        m = re.match(r"rId(\d+)$", rel.attrib.get("Id", ""))
        if m:
            max_id = max(max_id, int(m.group(1)))
    return f"rId{max_id + 1}"


def inject_wps_cell_images(xlsx_path, images, log=print):
    """把图片以 WPS 原生 DISPIMG 格式注入已保存的 xlsx 文件。

    images: [(img_id, image_file_path), ...]
    前提：xlsx 中对应单元格已写入公式 =_xlfn.DISPIMG("img_id",1)。
    该函数会：
      1. 把图片文件写入 xl/media/
      2. 创建/更新 xl/cellimages.xml 及其 rels
      3. 在 [Content_Types].xml 和 workbook.xml.rels 中登记 cellimages
      4. 为 DISPIMG 公式单元格补上缓存值 <v>（WPS 需要）
    """
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        files = {n: zin.read(n) for n in zin.namelist() if not n.endswith("/")}

    # ---- [Content_Types].xml ----
    ct_path = "[Content_Types].xml"
    ct_root = ET.fromstring(files[ct_path])

    def ensure_default(ext):
        ext = ext.lower().lstrip(".")
        for node in ct_root.findall(f"{{{_NS_CT}}}Default"):
            if node.attrib.get("Extension", "").lower() == ext:
                return
        ET.SubElement(ct_root, f"{{{_NS_CT}}}Default", {
            "Extension": ext,
            "ContentType": _EXT_CONTENT_TYPES.get('.' + ext, 'application/octet-stream')
        })

    for _, path in images:
        ensure_default(os.path.splitext(path)[1])

    has_override = any(
        node.attrib.get("PartName") == "/xl/cellimages.xml"
        for node in ct_root.findall(f"{{{_NS_CT}}}Override")
    )
    if not has_override:
        ET.SubElement(ct_root, f"{{{_NS_CT}}}Override", {
            "PartName": "/xl/cellimages.xml",
            "ContentType": WPS_CELLIMAGE_CONTENT_TYPE
        })
    files[ct_path] = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)

    # ---- workbook.xml.rels 登记 cellimages ----
    wb_rels_path = "xl/_rels/workbook.xml.rels"
    rel_root = ET.fromstring(files[wb_rels_path])
    if not any(rel.attrib.get("Target", "").replace("\\", "/").endswith("cellimages.xml")
               for rel in rel_root.findall(f"{{{_NS_REL}}}Relationship")):
        ET.SubElement(rel_root, f"{{{_NS_REL}}}Relationship", {
            "Id": _next_rid(rel_root),
            "Type": WPS_CELLIMAGE_REL_TYPE,
            "Target": "cellimages.xml"
        })
    files[wb_rels_path] = ET.tostring(rel_root, encoding="utf-8", xml_declaration=True)

    # ---- cellimages.xml 及其 rels ----
    ci_path = "xl/cellimages.xml"
    ci_rels_path = "xl/_rels/cellimages.xml.rels"
    if ci_path in files:
        ci_root = ET.fromstring(files[ci_path])
    else:
        ci_root = ET.Element(f"{{{_NS_ETC}}}cellImages")
    if ci_rels_path in files:
        ci_rel_root = ET.fromstring(files[ci_rels_path])
    else:
        ci_rel_root = ET.Element(f"{{{_NS_REL}}}Relationships")

    existing_media = {n for n in files if n.startswith("xl/media/")}
    seq = 0
    for img_id, img_path in images:
        ext = os.path.splitext(img_path)[1].lower()
        while True:
            seq += 1
            media_name = f"cellimage_{seq}{ext}"
            if f"xl/media/{media_name}" not in existing_media:
                break
        with open(img_path, "rb") as f:
            files[f"xl/media/{media_name}"] = f.read()

        rid = _next_rid(ci_rel_root)
        ET.SubElement(ci_rel_root, f"{{{_NS_REL}}}Relationship", {
            "Id": rid, "Type": _IMAGE_REL_TYPE, "Target": f"media/{media_name}"
        })

        # 图片尺寸（EMU）
        try:
            with PILImage.open(img_path) as im:
                w, h = im.size
        except Exception:
            w, h = 100, 100

        cell_image = ET.SubElement(ci_root, f"{{{_NS_ETC}}}cellImage")
        pic = ET.SubElement(cell_image, f"{{{_NS_XDR}}}pic")
        nv = ET.SubElement(pic, f"{{{_NS_XDR}}}nvPicPr")
        ET.SubElement(nv, f"{{{_NS_XDR}}}cNvPr", {
            "id": str(1000 + len(list(ci_root))),
            "name": img_id,
            "descr": os.path.splitext(os.path.basename(img_path))[0]
        })
        ET.SubElement(nv, f"{{{_NS_XDR}}}cNvPicPr")
        blip_fill = ET.SubElement(pic, f"{{{_NS_XDR}}}blipFill")
        ET.SubElement(blip_fill, f"{{{_NS_A}}}blip", {
            f"{{{_NS_R}}}embed": rid, "cstate": "print"
        })
        stretch = ET.SubElement(blip_fill, f"{{{_NS_A}}}stretch")
        ET.SubElement(stretch, f"{{{_NS_A}}}fillRect")
        sp_pr = ET.SubElement(pic, f"{{{_NS_XDR}}}spPr")
        xfrm = ET.SubElement(sp_pr, f"{{{_NS_A}}}xfrm")
        ET.SubElement(xfrm, f"{{{_NS_A}}}off", {"x": "0", "y": "0"})
        ET.SubElement(xfrm, f"{{{_NS_A}}}ext", {"cx": str(int(w * 9525)), "cy": str(int(h * 9525))})
        geom = ET.SubElement(sp_pr, f"{{{_NS_A}}}prstGeom", {"prst": "rect"})
        ET.SubElement(geom, f"{{{_NS_A}}}avLst")

    files[ci_path] = ET.tostring(ci_root, encoding="utf-8", xml_declaration=True)
    files[ci_rels_path] = ET.tostring(ci_rel_root, encoding="utf-8", xml_declaration=True)

    # ---- 为 DISPIMG 公式单元格补缓存值 <v> ----
    # openpyxl 保存的公式单元格只有 <f>，WPS 需要 <v>=DISPIMG("ID",1)</v>
    patched = 0
    for name in list(files):
        if not (name.startswith("xl/worksheets/") and name.endswith(".xml")):
            continue
        text = files[name].decode("utf-8")
        if "DISPIMG" not in text:
            continue
        def _add_cache(m):
            nonlocal patched
            patched += 1
            img_id = m.group(2)
            q = m.group(1)
            return (f'<f>_xlfn.DISPIMG({q}{img_id}{q},1)</f>'
                    f'<v>=DISPIMG({q}{img_id}{q},1)</v>')
        new_text = re.sub(
            r'<f>_xlfn\.DISPIMG\(("|&quot;)([^"&<]+)\1,1\)</f>(?:<v\s*/>|<v></v>|<v>[^<]*</v>)?',
            _add_cache, text)
        files[name] = new_text.encode("utf-8")

    log(f"已为 {patched} 个单元格写入 DISPIMG 缓存值。")

    # ---- 重写 zip ----
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    shutil.move(tmp, xlsx_path)


# ============== 核心逻辑 ==============

class ExcelCellImageEmbedder:
    def __init__(self, excel_path: str, image_folder: str,
                 match_column: str, target_column: str,
                 match_mode: str = "exact",
                 use_name_without_ext: bool = True,
                 start_row: int = 1,
                 cell_width_px: int = 120,
                 cell_height_px: int = 120,
                 fuzzy_threshold: float = 0.3,
                 embed_format: str = "wps",
                 log_callback=None):
        self.excel_path = excel_path
        self.image_folder = image_folder
        self.match_column = match_column.strip().upper()
        self.target_column = target_column.strip().upper()
        self.match_mode = match_mode
        self.use_name_without_ext = use_name_without_ext
        self.start_row = start_row
        self.cell_width_px = cell_width_px
        self.cell_height_px = cell_height_px
        self.fuzzy_threshold = fuzzy_threshold
        self.embed_format = embed_format  # "wps" | "excel365"
        self.log_callback = log_callback or print

        self.match_col_idx = column_letter_to_index(self.match_column)
        self.target_col_idx = column_letter_to_index(self.target_column)
        self.converter = ImageConverter()
        self._fmt_cache = {}

    def log(self, msg: str):
        self.log_callback(msg)

    # ---------- 匹配 ----------

    def _find_matches(self, ws, lookup):
        """返回 {行号: (img_path, score, matched_name, match_type)}。"""
        results = {}
        for row in range(self.start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=self.match_col_idx).value
            if cell_value is None:
                continue
            key = str(cell_value).strip()
            if not key:
                continue

            normalized_key = key.replace(' ', '_')
            img_path = lookup.get(key) or lookup.get(normalized_key)
            if img_path:
                results[row] = (img_path, 1.0,
                                key if key in lookup else normalized_key, "exact")
            elif self.match_mode != "exact":
                img_path, score, matched_name = find_best_fuzzy_match(
                    key, lookup, threshold=self.fuzzy_threshold
                )
                if img_path:
                    results[row] = (img_path, score, matched_name, "fuzzy")
        return results

    # ---------- 样式转换（openpyxl -> xlsxwriter） ----------

    def _style_to_format(self, wb_out, cell):
        """把 openpyxl 单元格样式尽量转换为 xlsxwriter format，带缓存。"""
        font = cell.font
        fill = cell.fill
        align = cell.alignment

        font_color = None
        try:
            if font.color and font.color.type == 'rgb' and isinstance(font.color.value, str):
                font_color = '#' + font.color.value[-6:]
        except Exception:
            pass

        bg_color = None
        try:
            if fill and fill.fill_type == 'solid':
                rgb = fill.start_color.value
                if isinstance(rgb, str):
                    bg_color = '#' + rgb[-6:]
        except Exception:
            pass

        key = (bool(font.bold), bool(font.italic), font.size, font.name,
               font_color, bg_color, align.horizontal, align.vertical,
               bool(align.wrap_text))
        if key in self._fmt_cache:
            return self._fmt_cache[key]

        props = {}
        if font.bold:
            props['bold'] = True
        if font.italic:
            props['italic'] = True
        if font.size:
            props['font_size'] = font.size
        if font.name:
            props['font_name'] = font.name
        if font_color and font_color.lower() != '#000000':
            props['font_color'] = font_color
        if bg_color and bg_color.lower() not in ('#ffffff', '#000000'):
            props['bg_color'] = bg_color
            props['pattern'] = 1
        h_map = {'center': 'center', 'left': 'left', 'right': 'right',
                 'centerContinuous': 'center'}
        if align.horizontal in h_map:
            props['align'] = h_map[align.horizontal]
        v_map = {'center': 'vcenter', 'top': 'top', 'bottom': 'bottom'}
        if align.vertical in v_map:
            props['valign'] = v_map[align.vertical]
        if align.wrap_text:
            props['text_wrap'] = True

        fmt = wb_out.add_format(props) if props else None
        self._fmt_cache[key] = fmt
        return fmt

    def _write_value(self, ws_out, row0, col0, value, fmt=None, wb_out=None):
        """按值类型写入 xlsxwriter 单元格（row0/col0 为 0-based）。"""
        if isinstance(value, datetime.datetime):
            dfmt = wb_out.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
            if fmt:
                # 样式与日期格式无法合并缓存时优先保证日期可读
                pass
            ws_out.write_datetime(row0, col0, value, dfmt)
        elif isinstance(value, datetime.date):
            dfmt = wb_out.add_format({'num_format': 'yyyy-mm-dd'})
            ws_out.write_datetime(row0, col0,
                                  datetime.datetime(value.year, value.month, value.day), dfmt)
        elif isinstance(value, str) and value.startswith('='):
            # openpyxl 读到的公式按文本写入，避免 #NAME 风险
            ws_out.write_string(row0, col0, value, fmt)
        else:
            ws_out.write(row0, col0, value, fmt)

    def _copy_sheet(self, ws_src, wb_out):
        """把一个 openpyxl 工作表的值/样式/合并单元格/行列尺寸复制到新工作簿。"""
        ws_out = wb_out.add_worksheet(ws_src.title[:31])

        # 列宽
        for letter, dim in ws_src.column_dimensions.items():
            if dim.width:
                try:
                    col0 = column_index_from_string(letter) - 1
                    ws_out.set_column(col0, col0, dim.width)
                except Exception:
                    pass
        # 行高
        for idx, dim in ws_src.row_dimensions.items():
            if dim.height:
                ws_out.set_row(idx - 1, min(dim.height, MAX_ROW_HEIGHT_PT))

        merged_topleft = {}
        for mr in ws_src.merged_cells.ranges:
            merged_topleft[(mr.min_row, mr.min_col)] = mr

        # 单元格值
        for row in ws_src.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                fmt = self._style_to_format(wb_out, c)
                self._write_value(ws_out, c.row - 1, c.column - 1, c.value, fmt, wb_out)

        # 合并单元格（覆盖重写左上角值）
        for (r, col), mr in merged_topleft.items():
            tl = ws_src.cell(row=r, column=col)
            fmt = self._style_to_format(wb_out, tl)
            try:
                ws_out.merge_range(mr.min_row - 1, mr.min_col - 1,
                                   mr.max_row - 1, mr.max_col - 1,
                                   tl.value if tl.value is not None else '', fmt)
            except Exception:
                pass
        return ws_out

    # ---------- 主流程 ----------

    def run(self):
        if not os.path.isfile(self.excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")
        if not os.path.isdir(self.image_folder):
            raise NotADirectoryError(f"图片文件夹不存在: {self.image_folder}")

        with_ext, without_ext = get_image_files(self.image_folder)
        self.log(f"图片文件夹中共找到 {len(with_ext)} 个图片文件。")

        wb_src = load_workbook(self.excel_path)
        ws_src = wb_src.active
        self.log(f"已打开工作簿，当前工作表: {ws_src.title}")
        if self.match_mode == "contains":
            self.log(f"模糊匹配阈值: {self.fuzzy_threshold:.0%}")

        lookup = without_ext if self.use_name_without_ext else with_ext
        matches = self._find_matches(ws_src, lookup)

        total = 0
        for row in range(self.start_row, ws_src.max_row + 1):
            v = ws_src.cell(row=row, column=self.match_col_idx).value
            if v is not None and str(v).strip():
                total += 1

        matched = len(matches)
        exact_matched = sum(1 for m in matches.values() if m[3] == "exact")
        fuzzy_matched = matched - exact_matched

        if self.embed_format == "wps":
            return self._run_wps(wb_src, ws_src, matches,
                                 total, exact_matched, fuzzy_matched)
        return self._run_excel365(wb_src, ws_src, matches,
                                  total, exact_matched, fuzzy_matched)

    def _run_excel365(self, wb_src, ws_src, matches,
                      total, exact_matched, fuzzy_matched):
        matched = len(matches)
        unmatched = total - matched

        output_path = self._get_output_path()
        wb_out = xlsxwriter.Workbook(output_path, {'in_memory': True})
        self._fmt_cache = {}

        # 目标列宽、匹配行行高
        target_col0 = self.target_col_idx - 1

        # 逐个 sheet 复制；当前工作表需要嵌入图片
        for ws in wb_src.worksheets:
            ws_out = self._copy_sheet(ws, wb_out)
            if ws is not ws_src:
                continue

            # 目标列宽调整到单元格图片宽度
            ws_out.set_column_pixels(target_col0, target_col0, self.cell_width_px)

            # 标注列：匹配类型、匹配图片名
            annotate_col_idx = max(ws.max_column, self.target_col_idx) + 1
            imgname_col_idx = annotate_col_idx + 1
            header_fmt = wb_out.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True})
            center_fmt = wb_out.add_format({'align': 'center', 'valign': 'vcenter'})
            left_fmt = wb_out.add_format({'align': 'left', 'valign': 'vcenter'})
            if self.start_row > 1:
                ws_out.write(self.start_row - 2, annotate_col_idx - 1, "匹配类型", header_fmt)
                ws_out.write(self.start_row - 2, imgname_col_idx - 1, "匹配图片名", header_fmt)
            ws_out.set_column(annotate_col_idx - 1, annotate_col_idx - 1, 22)
            ws_out.set_column(imgname_col_idx - 1, imgname_col_idx - 1, 40)

            # 嵌入图片
            for row in range(self.start_row, ws.max_row + 1):
                v = ws.cell(row=row, column=self.match_col_idx).value
                if v is None or not str(v).strip():
                    continue
                key = str(v).strip()

                if row in matches:
                    img_path, score, _name, mtype = matches[row]
                    rel_path = os.path.relpath(img_path, self.image_folder)

                    # 行高调整（与原有行高取较大值）
                    desired_pt = self.cell_height_px * PX_TO_PT
                    existing = ws.row_dimensions[row].height if row in ws.row_dimensions else None
                    new_h = max(existing or 0, desired_pt)
                    ws_out.set_row(row - 1, min(new_h, MAX_ROW_HEIGHT_PT))

                    # 计算等比缩放
                    embed_path = self.converter.get_embeddable_path(img_path)
                    size = self.converter.get_image_size(img_path)
                    opts = {'description': os.path.basename(img_path)}
                    if size:
                        w, h = size
                        scale = min(self.cell_width_px / w, self.cell_height_px / h)
                        opts['x_scale'] = scale
                        opts['y_scale'] = scale

                    try:
                        ws_out.embed_image(row - 1, target_col0, embed_path, opts)
                    except Exception as e:
                        self.log(f"[嵌入失败] 行 {row}: '{key}' -> {rel_path}: {e}")
                        unmatched += 1
                        matched -= 1
                        if mtype == "exact":
                            exact_matched -= 1
                        else:
                            fuzzy_matched -= 1
                        ws_out.write(row - 1, annotate_col_idx - 1, "嵌入失败", center_fmt)
                        continue

                    prefix = "[精确匹配]" if mtype == "exact" else "[模糊匹配]"
                    if mtype == "exact":
                        ws_out.write(row - 1, annotate_col_idx - 1, "精确匹配", center_fmt)
                    else:
                        ws_out.write(row - 1, annotate_col_idx - 1,
                                     "模糊匹配（相似度: {:.0%}）".format(score), center_fmt)
                    ws_out.write(row - 1, imgname_col_idx - 1,
                                 os.path.basename(img_path), left_fmt)
                    self.log(f"{prefix} 行 {row}: '{key}' -> {rel_path} (相似度: {score:.0%})")
                else:
                    ws_out.write(row - 1, annotate_col_idx - 1, "未匹配", center_fmt)
                    if self.match_mode == "contains":
                        self.log(f"[未匹配]   行 {row}: '{key}' 未找到相似度≥{self.fuzzy_threshold:.0%}的图片")
                    else:
                        self.log(f"[未匹配]   行 {row}: '{key}' 未找到对应图片")

        # 未匹配数据写入独立 sheet
        if unmatched > 0:
            self._write_unmatched_sheet(wb_out, ws_src, matches)

        wb_out.close()

        self.log(f"\n处理完成！")
        if self.match_mode == "contains" and matched > 0:
            self.log(f"总数据行: {total} | 精确匹配: {exact_matched} | 模糊匹配: {fuzzy_matched} | 未匹配: {unmatched}")
        else:
            self.log(f"总数据行: {total} | 成功匹配: {matched} | 未匹配: {unmatched}")
        self.log(f"结果已保存至: {output_path}")
        return output_path

    def _run_wps(self, wb, ws, matches, total, exact_matched, fuzzy_matched):
        """WPS 原生 DISPIMG 嵌入：openpyxl 直接改原工作簿 + zip 注入 cellimages。"""
        matched = len(matches)
        unmatched = total - matched

        # 标注列：匹配类型、匹配图片名
        annotate_col_idx = max(ws.max_column, self.target_col_idx) + 1
        imgname_col_idx = annotate_col_idx + 1
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        if self.start_row > 1:
            h = ws.cell(row=self.start_row - 1, column=annotate_col_idx, value="匹配类型")
            h.alignment = center
            h = ws.cell(row=self.start_row - 1, column=imgname_col_idx, value="匹配图片名")
            h.alignment = center
        ws.column_dimensions[get_column_letter(annotate_col_idx)].width = 22
        ws.column_dimensions[get_column_letter(imgname_col_idx)].width = 40

        # 目标列宽（字符单位 ≈ (px - 5) / 7）
        ws.column_dimensions[get_column_letter(self.target_col_idx)].width = \
            max((self.cell_width_px - 5) / 7, 2)

        pending = []  # [(img_id, embed_path)]
        embed_failed = 0

        for row in range(self.start_row, ws.max_row + 1):
            v = ws.cell(row=row, column=self.match_col_idx).value
            if v is None or not str(v).strip():
                continue
            key = str(v).strip()

            if row not in matches:
                ac = ws.cell(row=row, column=annotate_col_idx, value="未匹配")
                ac.alignment = center
                if self.match_mode == "contains":
                    self.log(f"[未匹配]   行 {row}: '{key}' 未找到相似度≥{self.fuzzy_threshold:.0%}的图片")
                else:
                    self.log(f"[未匹配]   行 {row}: '{key}' 未找到对应图片")
                continue

            img_path, score, _name, mtype = matches[row]
            rel_path = os.path.relpath(img_path, self.image_folder)

            try:
                embed_path = self.converter.get_embeddable_path(img_path)
            except Exception as e:
                self.log(f"[嵌入失败] 行 {row}: '{key}' -> {rel_path}: {e}")
                ac = ws.cell(row=row, column=annotate_col_idx, value="嵌入失败")
                ac.alignment = center
                embed_failed += 1
                continue

            img_id = "ID_" + secrets.token_hex(16).upper()
            cell = ws.cell(row=row, column=self.target_col_idx)
            cell.value = f'=_xlfn.DISPIMG("{img_id}",1)'
            cell.alignment = center
            pending.append((img_id, embed_path))

            # 行高调整（与原有行高取较大值）
            desired_pt = self.cell_height_px * PX_TO_PT
            existing = ws.row_dimensions[row].height
            ws.row_dimensions[row].height = min(max(existing or 0, desired_pt),
                                                MAX_ROW_HEIGHT_PT)

            prefix = "[精确匹配]" if mtype == "exact" else "[模糊匹配]"
            if mtype == "exact":
                ac = ws.cell(row=row, column=annotate_col_idx, value="精确匹配")
            else:
                ac = ws.cell(row=row, column=annotate_col_idx,
                             value="模糊匹配（相似度: {:.0%}）".format(score))
            ac.alignment = center
            nc = ws.cell(row=row, column=imgname_col_idx,
                         value=os.path.basename(img_path))
            nc.alignment = left
            self.log(f"{prefix} 行 {row}: '{key}' -> {rel_path} (相似度: {score:.0%})")

        matched -= embed_failed
        unmatched += embed_failed

        # 未匹配数据写入独立 sheet
        if unmatched > 0:
            self._write_unmatched_sheet_openpyxl(wb, ws, matches)

        output_path = self._get_output_path()
        wb.save(output_path)

        self.log("正在写入 WPS 嵌入图片数据（cellimages.xml）…")
        inject_wps_cell_images(output_path, pending, log=self.log)

        self.log(f"\n处理完成！（WPS 嵌入格式）")
        if self.match_mode == "contains" and matched > 0:
            self.log(f"总数据行: {total} | 精确匹配: {exact_matched} | 模糊匹配: {fuzzy_matched} | 未匹配: {unmatched}")
        else:
            self.log(f"总数据行: {total} | 成功匹配: {matched} | 未匹配: {unmatched}")
        self.log(f"结果已保存至: {output_path}")
        return output_path

    def _write_unmatched_sheet_openpyxl(self, wb, ws, matches):
        sheet_name = "未匹配数据"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_u = wb.create_sheet(title=sheet_name)

        max_col = ws.max_column
        headers = ["原始行号", "匹配键值"]
        if self.start_row > 1:
            for c in range(1, max_col + 1):
                headers.append(ws.cell(row=self.start_row - 1, column=c).value)
        else:
            for c in range(1, max_col + 1):
                headers.append(f"列{get_column_letter(c)}")
        for col_idx, h in enumerate(headers, start=1):
            ws_u.cell(row=1, column=col_idx, value=h)

        r = 2
        for row in range(self.start_row, ws.max_row + 1):
            v = ws.cell(row=row, column=self.match_col_idx).value
            if v is None or not str(v).strip():
                continue
            if row in matches:
                continue
            ws_u.cell(row=r, column=1, value=row)
            ws_u.cell(row=r, column=2, value=str(v).strip())
            for c in range(1, max_col + 1):
                val = ws.cell(row=row, column=c).value
                # DISPIMG 公式单元格不复制
                if isinstance(val, str) and "DISPIMG" in val:
                    continue
                ws_u.cell(row=r, column=c + 2, value=val)
            r += 1

        ws_u.column_dimensions['A'].width = 10
        ws_u.column_dimensions['B'].width = 30
        self.log(f"\n已创建 '{sheet_name}' sheet，共记录 {r - 2} 条未匹配数据。")

    def _write_unmatched_sheet(self, wb_out, ws_src, matches):
        ws_out = wb_out.add_worksheet("未匹配数据")
        bold = wb_out.add_format({'bold': True})

        max_col = ws_src.max_column
        headers = ["原始行号", "匹配键值"]
        if self.start_row > 1:
            for c in range(1, max_col + 1):
                headers.append(ws_src.cell(row=self.start_row - 1, column=c).value)
        else:
            for c in range(1, max_col + 1):
                headers.append(f"列{get_column_letter(c)}")

        for col0, h in enumerate(headers):
            ws_out.write(0, col0, h, bold)

        r = 1
        for row in range(self.start_row, ws_src.max_row + 1):
            v = ws_src.cell(row=row, column=self.match_col_idx).value
            if v is None or not str(v).strip():
                continue
            if row in matches:
                continue
            ws_out.write(r, 0, row)
            ws_out.write(r, 1, str(v).strip())
            for c in range(1, max_col + 1):
                val = ws_src.cell(row=row, column=c).value
                if val is not None:
                    self._write_value(ws_out, r, c + 1, val, None, wb_out)
            r += 1

        ws_out.set_column(0, 0, 10)
        ws_out.set_column(1, 1, 30)
        self.log(f"\n已创建 '未匹配数据' sheet，共记录 {r - 1} 条未匹配数据。")

    def _get_output_path(self):
        p = Path(self.excel_path)
        suffix = "_WPS嵌入图片" if self.embed_format == "wps" else "_365嵌入图片"
        return str(p.parent / f"{p.stem}{suffix}{p.suffix}")


# ============== GUI ==============

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 图片嵌入单元格工具（真嵌入）")
        self.geometry("720x700")
        self.minsize(700, 620)

        self._build_ui()
        self._set_defaults()
        self._toggle_threshold_options()

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        main = ttk.Frame(self, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        # ---- 文件选择区域 ----
        file_frame = ttk.LabelFrame(main, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(file_frame, text="Excel 文件:").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.entry_excel = ttk.Entry(file_frame, width=50)
        self.entry_excel.grid(row=0, column=1, sticky=tk.EW, padx=6)
        ttk.Button(file_frame, text="浏览…", command=self._browse_excel).grid(row=0, column=2)

        ttk.Label(file_frame, text="图片文件夹:").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.entry_folder = ttk.Entry(file_frame, width=50)
        self.entry_folder.grid(row=1, column=1, sticky=tk.EW, padx=6)
        ttk.Button(file_frame, text="浏览…", command=self._browse_folder).grid(row=1, column=2)

        file_frame.columnconfigure(1, weight=1)

        # ---- 参数配置区域 ----
        opts_frame = ttk.LabelFrame(main, text="匹配与嵌入参数", padding=10)
        opts_frame.pack(fill=tk.X, pady=(0, 10))

        # 第0行: 匹配列、目标列、起始行
        ttk.Label(opts_frame, text="匹配列 (如 A):").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.entry_match_col = ttk.Entry(opts_frame, width=8)
        self.entry_match_col.grid(row=0, column=1, sticky=tk.W, padx=6)

        ttk.Label(opts_frame, text="插入目标列 (如 D):").grid(row=0, column=2, sticky=tk.W, pady=4, padx=(16, 0))
        self.entry_target_col = ttk.Entry(opts_frame, width=8)
        self.entry_target_col.grid(row=0, column=3, sticky=tk.W, padx=6)

        ttk.Label(opts_frame, text="数据起始行:").grid(row=0, column=4, sticky=tk.W, pady=4, padx=(16, 0))
        self.entry_start_row = ttk.Entry(opts_frame, width=8)
        self.entry_start_row.grid(row=0, column=5, sticky=tk.W, padx=6)

        # 第1行: 匹配模式、模糊匹配阈值
        ttk.Label(opts_frame, text="匹配模式:").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.combo_match_mode = ttk.Combobox(
            opts_frame, values=["精确匹配", "模糊匹配"], width=10, state="readonly"
        )
        self.combo_match_mode.grid(row=1, column=1, sticky=tk.W, padx=6)
        self.combo_match_mode.bind("<<ComboboxSelected>>", lambda _e: self._toggle_threshold_options())

        self.frame_threshold = ttk.Frame(opts_frame)
        self.frame_threshold.grid(row=1, column=2, columnspan=2, sticky=tk.W, pady=4, padx=(16, 0))

        ttk.Label(self.frame_threshold, text="阈值:").pack(side=tk.LEFT)
        self.entry_threshold = ttk.Entry(self.frame_threshold, width=6)
        self.entry_threshold.pack(side=tk.LEFT, padx=4)
        ttk.Label(self.frame_threshold, text="(0~1)").pack(side=tk.LEFT)

        self.btn_help = tk.Button(
            self.frame_threshold, text="❓", width=2, height=1,
            font=("Arial", 10), cursor="hand2", relief=tk.FLAT, fg="#0066cc",
            command=self._show_threshold_help
        )
        self.btn_help.pack(side=tk.LEFT, padx=(6, 0))

        # 是否忽略扩展名
        self.var_ignore_ext = tk.BooleanVar(value=True)
        ttk.Checkbutton(opts_frame, text="匹配时忽略图片扩展名", variable=self.var_ignore_ext).grid(
            row=2, column=0, columnspan=3, sticky=tk.W, pady=4)

        # 嵌入格式
        ttk.Label(opts_frame, text="嵌入格式:").grid(row=2, column=3, sticky=tk.W, pady=4, padx=(16, 0))
        self.combo_embed_format = ttk.Combobox(
            opts_frame, values=["WPS（推荐）", "Excel 365"], width=12, state="readonly"
        )
        self.combo_embed_format.grid(row=2, column=4, columnspan=2, sticky=tk.W, padx=6)
        self.combo_embed_format.bind("<<ComboboxSelected>>", lambda _e: self._update_format_note())

        # 第3行: 单元格图片尺寸
        size_frame = ttk.Frame(opts_frame)
        size_frame.grid(row=3, column=0, columnspan=6, sticky=tk.W, pady=4)

        ttk.Label(size_frame, text="单元格宽度(px):").pack(side=tk.LEFT)
        self.entry_cell_w = ttk.Entry(size_frame, width=8)
        self.entry_cell_w.pack(side=tk.LEFT, padx=6)
        ttk.Label(size_frame, text="单元格高度(px):").pack(side=tk.LEFT, padx=(16, 0))
        self.entry_cell_h = ttk.Entry(size_frame, width=8)
        self.entry_cell_h.pack(side=tk.LEFT, padx=6)
        ttk.Label(size_frame, text="（行高列宽会自动调整到该尺寸）",
                  foreground="gray").pack(side=tk.LEFT, padx=(12, 0))

        # 兼容性提示
        self.note_label = tk.Label(opts_frame, text="", foreground="#b06000",
                                   justify=tk.LEFT, wraplength=660, anchor=tk.W)
        self.note_label.grid(row=4, column=0, columnspan=6, sticky=tk.W, pady=(6, 0))

        # ---- 操作按钮 ----
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        self.btn_run = ttk.Button(btn_frame, text="开始处理", command=self._on_run, width=16)
        self.btn_run.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="清空日志", command=self._clear_log, width=12).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="退出", command=self.destroy, width=10).pack(side=tk.LEFT)

        # ---- 日志区域 ----
        log_frame = ttk.LabelFrame(main, text="运行日志", padding=6)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.text_log = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, state=tk.DISABLED, height=12)
        self.text_log.pack(fill=tk.BOTH, expand=True)

        # 状态栏
        self.status = ttk.Label(main, text="就绪", anchor=tk.W, relief=tk.SUNKEN)
        self.status.pack(fill=tk.X, pady=(6, 0))

    def _set_defaults(self):
        self.entry_match_col.insert(0, "B")
        self.entry_target_col.insert(0, "D")
        self.entry_start_row.insert(0, "2")
        self.combo_match_mode.current(1)
        self.entry_threshold.insert(0, "0.3")
        self.entry_cell_w.insert(0, "120")
        self.entry_cell_h.insert(0, "120")
        self.combo_embed_format.current(0)
        self._update_format_note()

    def _update_format_note(self):
        if self.combo_embed_format.get().startswith("WPS"):
            self.note_label.configure(text=WPS_NOTE)
        else:
            self.note_label.configure(text=EXCEL365_NOTE)

    def _toggle_threshold_options(self):
        mode = self.combo_match_mode.get()
        if mode == "模糊匹配":
            self.entry_threshold.configure(state="normal")
        else:
            self.entry_threshold.configure(state="disabled")

    def _show_threshold_help(self):
        help_window = tk.Toplevel(self)
        help_window.title("模糊匹配阈值说明")
        help_window.geometry("520x520")
        help_window.transient(self)
        help_window.grab_set()

        text = tk.Text(help_window, wrap=tk.WORD, padx=16, pady=16, font=("Menlo", 11))
        text.pack(fill=tk.BOTH, expand=True)
        text.insert(tk.END, THRESHOLD_HELP_TEXT)
        text.configure(state=tk.DISABLED)

        ttk.Button(help_window, text="知道了", command=help_window.destroy).pack(pady=10)

        help_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - help_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - help_window.winfo_height()) // 2
        help_window.geometry(f"+{x}+{y}")

    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
        )
        if path:
            self.entry_excel.delete(0, tk.END)
            self.entry_excel.insert(0, path)

    def _browse_folder(self):
        path = filedialog.askdirectory(title="选择图片文件夹")
        if path:
            self.entry_folder.delete(0, tk.END)
            self.entry_folder.insert(0, path)

    def _log(self, msg: str):
        self.text_log.configure(state=tk.NORMAL)
        self.text_log.insert(tk.END, msg + "\n")
        self.text_log.see(tk.END)
        self.text_log.configure(state=tk.DISABLED)
        self.update_idletasks()

    def _clear_log(self):
        self.text_log.configure(state=tk.NORMAL)
        self.text_log.delete(1.0, tk.END)
        self.text_log.configure(state=tk.DISABLED)

    def _on_run(self):
        excel_path = self.entry_excel.get().strip()
        folder = self.entry_folder.get().strip()
        match_col = self.entry_match_col.get().strip()
        target_col = self.entry_target_col.get().strip()
        start_row_str = self.entry_start_row.get().strip()
        cell_w_str = self.entry_cell_w.get().strip()
        cell_h_str = self.entry_cell_h.get().strip()
        threshold_str = self.entry_threshold.get().strip()
        match_mode = "exact" if self.combo_match_mode.get() == "精确匹配" else "contains"
        ignore_ext = self.var_ignore_ext.get()
        embed_format = "wps" if self.combo_embed_format.get().startswith("WPS") else "excel365"

        if not excel_path or not os.path.isfile(excel_path):
            messagebox.showerror("错误", "请选择一个有效的 Excel 文件。")
            return
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("错误", "请选择一个有效的图片文件夹。")
            return
        if not match_col:
            messagebox.showerror("错误", "请填写匹配列。")
            return
        if not target_col:
            messagebox.showerror("错误", "请填写目标插入列。")
            return
        if match_col.upper() == target_col.upper():
            messagebox.showerror("错误", "匹配列和目标列不能相同。")
            return

        try:
            start_row = int(start_row_str)
            if start_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "起始行必须是正整数。")
            return

        try:
            cell_w = int(cell_w_str) if cell_w_str else 120
            cell_h = int(cell_h_str) if cell_h_str else 120
            if cell_w < 1 or cell_h < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "单元格宽度和高度必须是正整数。")
            return

        if cell_h * PX_TO_PT > MAX_ROW_HEIGHT_PT:
            messagebox.showerror(
                "错误",
                f"单元格高度过大：Excel 行高上限约 546 像素（409.5 磅），"
                f"当前设置 {cell_h} 像素超出限制。")
            return

        fuzzy_threshold = 0.3
        if match_mode == "contains":
            try:
                fuzzy_threshold = float(threshold_str)
                if not (0.0 <= fuzzy_threshold <= 1.0):
                    raise ValueError
            except ValueError:
                messagebox.showerror("错误", "模糊匹配阈值必须是 0~1 之间的小数（如 0.3）。")
                return

        self.btn_run.configure(state=tk.DISABLED)
        self.status.configure(text="正在处理，请稍候…")
        self._clear_log()
        self._log("开始处理…（真嵌入单元格模式）")
        self._log(f"嵌入格式: {'WPS (DISPIMG)' if embed_format == 'wps' else 'Excel 365 (richData)'}")
        self._log(f"Excel 文件: {excel_path}")
        self._log(f"图片文件夹: {folder}")
        self._log(f"匹配列: {match_col.upper()} -> 目标列: {target_col.upper()}, 起始行: {start_row}")
        self._log(f"匹配模式: {match_mode}" + (f", 阈值: {fuzzy_threshold:.0%}" if match_mode == "contains" else ""))
        self._log(f"忽略扩展名: {ignore_ext}")
        self._log(f"单元格尺寸: {cell_w} x {cell_h} px")
        self._log("-" * 50)

        def task():
            try:
                embedder = ExcelCellImageEmbedder(
                    excel_path=excel_path,
                    image_folder=folder,
                    match_column=match_col,
                    target_column=target_col,
                    match_mode=match_mode,
                    use_name_without_ext=ignore_ext,
                    start_row=start_row,
                    cell_width_px=cell_w,
                    cell_height_px=cell_h,
                    fuzzy_threshold=fuzzy_threshold,
                    embed_format=embed_format,
                    log_callback=self._log
                )
                out_path = embedder.run()
                self.status.configure(text=f"完成: {out_path}")
                if embed_format == "wps":
                    tip = "注意：该文件需用 WPS 打开查看；Microsoft Excel 中无法显示图片。"
                else:
                    tip = "注意：需要 Microsoft 365 或支持该格式的新版 WPS 打开才能正常显示。"
                messagebox.showinfo(
                    "完成",
                    f"处理完成！图片已真正嵌入单元格。\n\n{tip}\n\n结果已保存至:\n{out_path}")
            except Exception as e:
                self._log(f"\n发生错误: {e}")
                self.status.configure(text="处理出错")
                messagebox.showerror("错误", str(e))
            finally:
                self.btn_run.configure(state=tk.NORMAL)

        self.after(100, task)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
