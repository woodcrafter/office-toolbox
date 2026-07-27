#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 图片插入工具
功能：根据 Excel 指定列的单元格内容匹配图片文件名，将图片嵌入单元格或插入图片链接。
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# ============== 阈值说明文本 ==============

THRESHOLD_HELP_TEXT = """【阈值说明】

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【模糊匹配处理流程】

当选择"模糊匹配"时，系统会按以下两步处理：

  第1步 — 精确匹配
    先用 Excel 单元格值直接查找图片文件名。
    如果找到，直接记录为【精确匹配成功】，不再进行模糊匹配。

  第2步 — 模糊匹配（仅第1步未匹配时执行）
    对所有图片文件名逐一评分，选择分数最高的一个。
    如果最高分 ≥ 阈值，则记录为【模糊匹配成功】。

  第3步 — 未匹配
    前两步都找不到对应图片，记录为【未匹配】。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【评分等级】

100% — 精确匹配
   Excel 单元格值与图片文件名完全一致
   （或仅空格 ↔ 下划线差异）
   误匹配风险：无

85%  — 长字符串包含匹配
   Excel 值被完整包含在图片名中，且双方长度均≥10字符
   例如：Excel="DRA343"  图片="NO.JA-QS-0403-3732_V1.0-DRA343.png"
   误匹配风险：极低

60%  — 短字符串包含匹配
   Excel 值被包含在图片名中，但某一方长度<10字符
   例如：Excel="C001"     图片="NO.JA-QS-0403-0001_V1.1-C001.png"
   误匹配风险：中等（可能出现多个图片含相同短字符串）

<60% — LCS（最长公共子串）相似度
   两个字符串有较长的公共部分，但互不为子串
   例如：Excel="ABC-123"  图片="ABC-124"
   误匹配风险：随分数降低而升高

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【推荐阈值】

90%（0.90）— 极严格
   只接受精确匹配 + 长包含匹配。误匹配极少。

60%（0.60）— 较严格（推荐）
   接受精确匹配 + 包含匹配。平衡覆盖面与准确性。

30%（0.30）— 宽松（默认）
   接受部分相似匹配。覆盖面最广，可能有少量误匹配。

10%（0.10）— 极宽松
   几乎任何有点相似的都会匹配。误匹配风险较高，
   仅建议在人工复核后使用。
"""


# ============== 工具函数 ==============

def column_letter_to_index(col_letter: str) -> int:
    """将列字母（如 'A', 'AB'）转换为 1-based 列索引。"""
    col_letter = col_letter.strip().upper()
    if not col_letter:
        raise ValueError("列名不能为空")
    try:
        idx = 0
        for ch in col_letter:
            if 'A' <= ch <= 'Z':
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            else:
                raise ValueError
        return idx
    except Exception:
        raise ValueError(f"无效的列名: {col_letter}")


def index_to_column_letter(idx: int) -> str:
    """将 1-based 列索引转换为列字母。"""
    return get_column_letter(idx)


def get_image_files(folder: str):
    """递归获取文件夹及其所有子文件夹中支持的图片文件，
    返回 {文件名(无扩展名): 完整路径} 和 {文件名(含扩展名): 完整路径}。"""
    exts = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
    without_ext = {}
    with_ext = {}
    folder = os.path.abspath(folder)

    for root, _dirs, files in os.walk(folder):
        for f in files:
            p = os.path.join(root, f)
            name_lower = f.lower()
            ext = os.path.splitext(name_lower)[1]
            if ext in exts:
                if f in with_ext:
                    pass
                else:
                    with_ext[f] = p
                    without_ext[os.path.splitext(f)[0]] = p
    return with_ext, without_ext


def similarity_score(a: str, b: str) -> float:
    """
    计算两个字符串的相似度分数 (0.0 ~ 1.0)。
    使用最长公共子串(LCS)长度与较长字符串长度的比值。
    """
    if not a or not b:
        return 0.0
    a, b = a.lower(), b.lower()
    m, n = len(a), len(b)
    max_len = 0
    dp = [0] * (n + 1)
    for i in range(1, m + 1):
        prev = 0
        for j in range(1, n + 1):
            temp = dp[j]
            if a[i - 1] == b[j - 1]:
                dp[j] = prev + 1
                max_len = max(max_len, dp[j])
            else:
                dp[j] = 0
            prev = temp
    longer = max(m, n)
    return max_len / longer if longer > 0 else 0.0


def find_best_fuzzy_match(key: str, candidates: dict, threshold: float = 0.3) -> tuple:
    """
    在候选图片中找到与 key 最相似的一个。
    返回 (img_path, score, matched_name) 或 (None, 0.0, None)。

    评分策略（按优先级降序）：
    1. 精确匹配（含规范化后的精确匹配）    → score = 1.0
    2. 长字符串双向包含（长度均≥10）        → score = 0.85
    3. 短字符串双向包含                      → score = 0.60
    4. 最长公共子串(LCS)相似度              → score = similarity
    """
    normalized_key = key.replace(' ', '_')
    best_path = None
    best_name = None
    best_score = 0.0

    for name, path in candidates.items():
        score = 0.0

        # 策略1: 精确匹配（最高分）
        if key == name or normalized_key == name:
            score = 1.0
        # 策略2: 双向包含
        elif key in name or name in key or normalized_key in name or name in normalized_key:
            if len(key) >= 10 and len(name) >= 10:
                score = 0.85
            else:
                score = 0.60
        # 策略3: LCS 相似度评分
        else:
            score = similarity_score(key, name)

        if score > best_score:
            best_score = score
            best_path = path
            best_name = name

    if best_score >= threshold:
        return best_path, best_score, best_name
    return None, 0.0, None


# ============== 核心逻辑 ==============

class ExcelImageInserter:
    def __init__(self, excel_path: str, image_folder: str,
                 match_column: str, target_column: str,
                 match_mode: str = "exact",
                 use_name_without_ext: bool = True,
                 start_row: int = 1,
                 image_width: int = 100,
                 image_height: int = 100,
                 insert_mode: str = "embed",
                 link_display_text: str = "查看图片",
                 fuzzy_threshold: float = 0.3,
                 log_callback=None):
        self.excel_path = excel_path
        self.image_folder = image_folder
        self.match_column = match_column.strip().upper()
        self.target_column = target_column.strip().upper()
        self.match_mode = match_mode
        self.use_name_without_ext = use_name_without_ext
        self.start_row = start_row
        self.image_width = image_width
        self.image_height = image_height
        self.insert_mode = insert_mode
        self.link_display_text = link_display_text
        self.fuzzy_threshold = fuzzy_threshold
        self.log_callback = log_callback or print

        self.match_col_idx = column_letter_to_index(self.match_column)
        self.target_col_idx = column_letter_to_index(self.target_column)

    def log(self, msg: str):
        self.log_callback(msg)

    def run(self):
        if not os.path.isfile(self.excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")
        if not os.path.isdir(self.image_folder):
            raise NotADirectoryError(f"图片文件夹不存在: {self.image_folder}")

        with_ext, without_ext = get_image_files(self.image_folder)
        self.log(f"图片文件夹中共找到 {len(with_ext)} 个图片文件。")

        wb = load_workbook(self.excel_path)
        ws = wb.active
        self.log(f"已打开工作簿，当前工作表: {ws.title}")
        self.log(f"插入模式: {'嵌入图片' if self.insert_mode == 'embed' else '插入超链接'}")
        if self.match_mode == "contains":
            self.log(f"模糊匹配阈值: {self.fuzzy_threshold:.0%}")

        matched = 0
        exact_matched = 0
        fuzzy_matched = 0
        unmatched = 0
        total = 0
        unmatched_rows = []

        max_row = ws.max_row
        max_col = ws.max_column
        lookup = without_ext if self.use_name_without_ext else with_ext

        # 在现有数据（含目标插入列）后面新增两列："匹配类型"、"匹配图片名"
        annotate_col_idx = max(max_col, self.target_col_idx) + 1
        imgname_col_idx = annotate_col_idx + 1
        if self.start_row > 1:
            header_cell = ws.cell(row=self.start_row - 1, column=annotate_col_idx, value="匹配类型")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            imgname_header = ws.cell(row=self.start_row - 1, column=imgname_col_idx, value="匹配图片名")
            imgname_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(annotate_col_idx)].width = 22
        ws.column_dimensions[get_column_letter(imgname_col_idx)].width = 40

        for row in range(self.start_row, max_row + 1):
            cell = ws.cell(row=row, column=self.match_col_idx)
            cell_value = cell.value
            if cell_value is None:
                continue
            total += 1
            key = str(cell_value).strip()
            if not key:
                continue

            normalized_key = key.replace(' ', '_')
            img_path = None
            match_score = 0.0
            matched_name = None
            match_type = None  # "exact" | "fuzzy" | None

            if self.match_mode == "exact":
                # 纯精确匹配模式
                img_path = lookup.get(key) or lookup.get(normalized_key)
                if img_path:
                    match_score = 1.0
                    matched_name = key if key in lookup else normalized_key
                    match_type = "exact"
            else:
                # 模糊匹配模式：先精确匹配，失败后再模糊匹配
                # 第1步：精确匹配
                img_path = lookup.get(key) or lookup.get(normalized_key)
                if img_path:
                    match_score = 1.0
                    matched_name = key if key in lookup else normalized_key
                    match_type = "exact"
                else:
                    # 第2步：模糊匹配
                    img_path, match_score, matched_name = find_best_fuzzy_match(
                        key, lookup, threshold=self.fuzzy_threshold
                    )
                    if img_path:
                        match_type = "fuzzy"

            if img_path:
                target_cell = ws.cell(row=row, column=self.target_col_idx)
                rel_path = os.path.relpath(img_path, self.image_folder)

                # 根据匹配类型选择日志前缀并写入标注列
                annotate_cell = ws.cell(row=row, column=annotate_col_idx)
                if match_type == "exact":
                    prefix = "[精确匹配]"
                    annotate_cell.value = "精确匹配"
                    exact_matched += 1
                else:
                    prefix = "[模糊匹配]"
                    annotate_cell.value = "模糊匹配（相似度: {:.0%}）".format(match_score)
                    fuzzy_matched += 1
                annotate_cell.alignment = Alignment(horizontal="center", vertical="center")
                # 写入匹配到的图片文件名（含扩展名）
                imgname_cell = ws.cell(row=row, column=imgname_col_idx, value=os.path.basename(img_path))
                imgname_cell.alignment = Alignment(horizontal="left", vertical="center")
                matched += 1

                if self.insert_mode == "embed":
                    img = OpenpyxlImage(img_path)
                    if self.image_width and self.image_height:
                        img.width = self.image_width
                        img.height = self.image_height
                    ws.add_image(img, f"{self.target_column}{row}")
                    self.log(f"{prefix} 行 {row}: '{key}' -> {rel_path} (相似度: {match_score:.0%})")

                else:  # link
                    target_cell.hyperlink = img_path
                    display_text = self.link_display_text or os.path.basename(img_path)
                    target_cell.value = display_text
                    target_cell.style = "Hyperlink"
                    target_cell.alignment = Alignment(horizontal="center", vertical="center")
                    self.log(f"{prefix} 行 {row}: '{key}' -> {rel_path} (相似度: {match_score:.0%})")

            else:
                unmatched += 1
                annotate_cell = ws.cell(row=row, column=annotate_col_idx, value="未匹配")
                annotate_cell.alignment = Alignment(horizontal="center", vertical="center")
                if self.match_mode == "contains":
                    self.log(f"[未匹配]   行 {row}: '{key}' 未找到相似度≥{self.fuzzy_threshold:.0%}的图片")
                else:
                    self.log(f"[未匹配]   行 {row}: '{key}' 未找到对应图片")
                row_data = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
                unmatched_rows.append((row, key, row_data))

        # 未匹配数据写入独立 sheet
        if unmatched_rows:
            sheet_name = "未匹配数据"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_unmatched = wb.create_sheet(title=sheet_name)

            headers = ["原始行号", "匹配键值"]
            if self.start_row > 1:
                for c in range(1, max_col + 1):
                    headers.append(ws.cell(row=self.start_row - 1, column=c).value)
            else:
                for c in range(1, max_col + 1):
                    headers.append(f"列{index_to_column_letter(c)}")

            for col_idx, h in enumerate(headers, start=1):
                ws_unmatched.cell(row=1, column=col_idx, value=h)

            for r_idx, (orig_row, key, row_data) in enumerate(unmatched_rows, start=2):
                ws_unmatched.cell(row=r_idx, column=1, value=orig_row)
                ws_unmatched.cell(row=r_idx, column=2, value=key)
                for c_idx, val in enumerate(row_data, start=3):
                    ws_unmatched.cell(row=r_idx, column=c_idx, value=val)

            for col in ws_unmatched.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_unmatched.column_dimensions[column].width = adjusted_width

            self.log(f"\n已创建 '{sheet_name}' sheet，共记录 {len(unmatched_rows)} 条未匹配数据。")

        output_path = self._get_output_path()
        wb.save(output_path)
        self.log(f"\n处理完成！")
        if self.match_mode == "contains" and matched > 0:
            self.log(f"总数据行: {total} | 精确匹配: {exact_matched} | 模糊匹配: {fuzzy_matched} | 未匹配: {unmatched}")
        else:
            self.log(f"总数据行: {total} | 成功匹配: {matched} | 未匹配: {unmatched}")
        self.log(f"结果已保存至: {output_path}")
        return output_path

    def _get_output_path(self):
        p = Path(self.excel_path)
        suffix_map = {"embed": "_插入图片", "link": "_插入链接"}
        suffix = suffix_map.get(self.insert_mode, "_处理结果")
        return str(p.parent / f"{p.stem}{suffix}{p.suffix}")


# ============== GUI ==============

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 图片批量插入工具")
        self.geometry("720x660")
        self.minsize(700, 580)

        self._build_ui()
        self._set_defaults()
        self._toggle_link_options()
        self._toggle_threshold_options()

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        main = ttk.Frame(self, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        # ---- 文件选择区域 ----
        file_frame = ttk.LabelFrame(main, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(file_frame, text="Excel 文件:").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.entry_excel = ttk.Entry(file_frame, width=50)
        self.entry_excel.grid(row=0, column=1, sticky=tk.EW, padx=6)
        ttk.Button(file_frame, text="浏览…", command=self._browse_excel).grid(row=0, column=2)

        ttk.Label(file_frame, text="图片文件夹:").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.entry_folder = ttk.Entry(file_frame, width=50)
        self.entry_folder.grid(row=1, column=1, sticky=tk.EW, padx=6)
        ttk.Button(file_frame, text="浏览…", command=self._browse_folder).grid(row=1, column=2)

        file_frame.columnconfigure(1, weight=1)

        # ---- 参数配置区域 ----
        opts_frame = ttk.LabelFrame(main, text="匹配与插入参数", padding=10)
        opts_frame.pack(fill=tk.X, pady=(0, 10))

        # 第0行: 匹配列、目标列、起始行
        ttk.Label(opts_frame, text="匹配列 (如 A):").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.entry_match_col = ttk.Entry(opts_frame, width=8)
        self.entry_match_col.grid(row=0, column=1, sticky=tk.W, padx=6)

        ttk.Label(opts_frame, text="插入目标列 (如 B):").grid(row=0, column=2, sticky=tk.W, pady=4, padx=(16, 0))
        self.entry_target_col = ttk.Entry(opts_frame, width=8)
        self.entry_target_col.grid(row=0, column=3, sticky=tk.W, padx=6)

        ttk.Label(opts_frame, text="数据起始行:").grid(row=0, column=4, sticky=tk.W, pady=4, padx=(16, 0))
        self.entry_start_row = ttk.Entry(opts_frame, width=8)
        self.entry_start_row.grid(row=0, column=5, sticky=tk.W, padx=6)

        # 第1行: 匹配模式、模糊匹配阈值、插入模式
        ttk.Label(opts_frame, text="匹配模式:").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.combo_match_mode = ttk.Combobox(
            opts_frame, values=["精确匹配", "模糊匹配"], width=10, state="readonly"
        )
        self.combo_match_mode.grid(row=1, column=1, sticky=tk.W, padx=6)
        self.combo_match_mode.bind("<<ComboboxSelected>>", lambda _e: self._toggle_threshold_options())

        # 模糊匹配阈值（含帮助按钮）
        self.frame_threshold = ttk.Frame(opts_frame)
        self.frame_threshold.grid(row=1, column=2, columnspan=2, sticky=tk.W, pady=4, padx=(16, 0))

        ttk.Label(self.frame_threshold, text="阈值:").pack(side=tk.LEFT)
        self.entry_threshold = ttk.Entry(self.frame_threshold, width=6)
        self.entry_threshold.pack(side=tk.LEFT, padx=4)
        ttk.Label(self.frame_threshold, text="(0~1)").pack(side=tk.LEFT)

        self.btn_help = tk.Button(
            self.frame_threshold,
            text="❓",
            width=2,
            height=1,
            font=("Arial", 10),
            cursor="hand2",
            relief=tk.FLAT,
            fg="#0066cc",
            command=self._show_threshold_help
        )
        self.btn_help.pack(side=tk.LEFT, padx=(6, 0))

        # 插入方式
        ttk.Label(opts_frame, text="插入方式:").grid(row=2, column=0, sticky=tk.W, pady=4)
        self.combo_insert_mode = ttk.Combobox(
            opts_frame,
            values=["嵌入图片", "插入超链接"],
            width=10,
            state="readonly"
        )
        self.combo_insert_mode.grid(row=2, column=1, sticky=tk.W, padx=6)
        self.combo_insert_mode.bind("<<ComboboxSelected>>", lambda _e: self._toggle_link_options())

        # 是否忽略扩展名
        self.var_ignore_ext = tk.BooleanVar(value=True)
        ttk.Checkbutton(opts_frame, text="匹配时忽略图片扩展名", variable=self.var_ignore_ext).grid(
            row=2, column=2, columnspan=3, sticky=tk.W, pady=4, padx=(16, 0))

        # 第3行: 图片尺寸（仅嵌入模式有效）
        self.frame_embed_opts = ttk.Frame(opts_frame)
        self.frame_embed_opts.grid(row=3, column=0, columnspan=6, sticky=tk.W, pady=4)

        ttk.Label(self.frame_embed_opts, text="图片宽度(px):").pack(side=tk.LEFT)
        self.entry_img_w = ttk.Entry(self.frame_embed_opts, width=8)
        self.entry_img_w.pack(side=tk.LEFT, padx=6)
        ttk.Label(self.frame_embed_opts, text="图片高度(px):").pack(side=tk.LEFT, padx=(16, 0))
        self.entry_img_h = ttk.Entry(self.frame_embed_opts, width=8)
        self.entry_img_h.pack(side=tk.LEFT, padx=6)

        # 第4行: 链接显示文字（仅链接模式有效）
        self.frame_link_opts = ttk.Frame(opts_frame)
        self.frame_link_opts.grid(row=4, column=0, columnspan=6, sticky=tk.W, pady=4)

        ttk.Label(self.frame_link_opts, text="链接显示文字:").pack(side=tk.LEFT)
        self.entry_link_text = ttk.Entry(self.frame_link_opts, width=30)
        self.entry_link_text.pack(side=tk.LEFT, padx=6)
        ttk.Label(self.frame_link_opts, text="（留空则显示图片文件名）", foreground="gray").pack(side=tk.LEFT)

        # ---- 操作按钮 ----
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        self.btn_run = ttk.Button(btn_frame, text="开始处理", command=self._on_run, width=16)
        self.btn_run.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="清空日志", command=self._clear_log, width=12).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="退出", command=self.destroy, width=10).pack(side=tk.LEFT)

        # ---- 日志区域 ----
        log_frame = ttk.LabelFrame(main, text="运行日志", padding=6)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.text_log = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, state=tk.DISABLED, height=12)
        self.text_log.pack(fill=tk.BOTH, expand=True)

        # 状态栏
        self.status = ttk.Label(main, text="就绪", anchor=tk.W, relief=tk.SUNKEN)
        self.status.pack(fill=tk.X, pady=(6, 0))

    def _set_defaults(self):
        self.entry_match_col.insert(0, "A")
        self.entry_target_col.insert(0, "B")
        self.entry_start_row.insert(0, "1")
        self.combo_match_mode.current(0)
        self.entry_threshold.insert(0, "0.3")
        self.combo_insert_mode.current(0)
        self.entry_img_w.insert(0, "100")
        self.entry_img_h.insert(0, "100")
        self.entry_link_text.insert(0, "查看图片")

    def _toggle_threshold_options(self):
        mode = self.combo_match_mode.get()
        if mode == "模糊匹配":
            self.entry_threshold.configure(state="normal")
        else:
            self.entry_threshold.configure(state="disabled")

    def _toggle_link_options(self):
        mode = self.combo_insert_mode.get()
        if mode == "插入超链接":
            self.frame_embed_opts.grid_remove()
            self.frame_link_opts.grid()
        else:
            self.frame_link_opts.grid_remove()
            self.frame_embed_opts.grid()

    def _show_threshold_help(self):
        help_window = tk.Toplevel(self)
        help_window.title("模糊匹配阈值说明")
        help_window.geometry("520x520")
        help_window.transient(self)
        help_window.grab_set()

        text = tk.Text(help_window, wrap=tk.WORD, padx=16, pady=16, font=("Menlo", 11))
        text.pack(fill=tk.BOTH, expand=True)
        text.insert(tk.END, THRESHOLD_HELP_TEXT)
        text.configure(state=tk.DISABLED)

        ttk.Button(help_window, text="知道了", command=help_window.destroy).pack(pady=10)

        help_window.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - help_window.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - help_window.winfo_height()) // 2
        help_window.geometry(f"+{x}+{y}")

    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
        )
        if path:
            self.entry_excel.delete(0, tk.END)
            self.entry_excel.insert(0, path)

    def _browse_folder(self):
        path = filedialog.askdirectory(title="选择图片文件夹")
        if path:
            self.entry_folder.delete(0, tk.END)
            self.entry_folder.insert(0, path)

    def _log(self, msg: str):
        self.text_log.configure(state=tk.NORMAL)
        self.text_log.insert(tk.END, msg + "\n")
        self.text_log.see(tk.END)
        self.text_log.configure(state=tk.DISABLED)
        self.update_idletasks()

    def _clear_log(self):
        self.text_log.configure(state=tk.NORMAL)
        self.text_log.delete(1.0, tk.END)
        self.text_log.configure(state=tk.DISABLED)

    def _on_run(self):
        excel_path = self.entry_excel.get().strip()
        folder = self.entry_folder.get().strip()
        match_col = self.entry_match_col.get().strip()
        target_col = self.entry_target_col.get().strip()
        start_row_str = self.entry_start_row.get().strip()
        img_w_str = self.entry_img_w.get().strip()
        img_h_str = self.entry_img_h.get().strip()
        threshold_str = self.entry_threshold.get().strip()
        match_mode = "exact" if self.combo_match_mode.get() == "精确匹配" else "contains"
        ignore_ext = self.var_ignore_ext.get()
        insert_mode = "embed" if self.combo_insert_mode.get() == "嵌入图片" else "link"
        link_text = self.entry_link_text.get().strip() if insert_mode == "link" else ""

        if not excel_path or not os.path.isfile(excel_path):
            messagebox.showerror("错误", "请选择一个有效的 Excel 文件。")
            return
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("错误", "请选择一个有效的图片文件夹。")
            return
        if not match_col:
            messagebox.showerror("错误", "请填写匹配列。")
            return
        if not target_col:
            messagebox.showerror("错误", "请填写目标插入列。")
            return
        if match_col.upper() == target_col.upper():
            messagebox.showerror("错误", "匹配列和目标列不能相同。")
            return

        try:
            start_row = int(start_row_str)
            if start_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "起始行必须是正整数。")
            return

        try:
            img_w = int(img_w_str) if img_w_str else 100
            img_h = int(img_h_str) if img_h_str else 100
            if img_w < 1 or img_h < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "图片宽度和高度必须是正整数。")
            return

        fuzzy_threshold = 0.3
        if match_mode == "contains":
            try:
                fuzzy_threshold = float(threshold_str)
                if not (0.0 <= fuzzy_threshold <= 1.0):
                    raise ValueError
            except ValueError:
                messagebox.showerror("错误", "模糊匹配阈值必须是 0~1 之间的小数（如 0.3）。")
                return

        self.btn_run.configure(state=tk.DISABLED)
        self.status.configure(text="正在处理，请稍候…")
        self._clear_log()
        self._log("开始处理…")
        self._log(f"Excel 文件: {excel_path}")
        self._log(f"图片文件夹: {folder}")
        self._log(f"匹配列: {match_col.upper()} -> 目标列: {target_col.upper()}, 起始行: {start_row}")
        self._log(f"匹配模式: {match_mode}" + (f", 阈值: {fuzzy_threshold:.0%}" if match_mode == "contains" else ""))
        self._log(f"忽略扩展名: {ignore_ext}")
        self._log(f"插入方式: {'嵌入图片' if insert_mode == 'embed' else '插入超链接'}")
        self._log("-" * 50)

        def task():
            try:
                inserter = ExcelImageInserter(
                    excel_path=excel_path,
                    image_folder=folder,
                    match_column=match_col,
                    target_column=target_col,
                    match_mode=match_mode,
                    use_name_without_ext=ignore_ext,
                    start_row=start_row,
                    image_width=img_w,
                    image_height=img_h,
                    insert_mode=insert_mode,
                    link_display_text=link_text,
                    fuzzy_threshold=fuzzy_threshold,
                    log_callback=self._log
                )
                out_path = inserter.run()
                self.status.configure(text=f"完成: {out_path}")
                mode_name = "嵌入图片" if insert_mode == "embed" else "插入链接"
                messagebox.showinfo("完成", f"{mode_name}处理完成！\n结果已保存至:\n{out_path}")
            except Exception as e:
                self._log(f"\n发生错误: {e}")
                self.status.configure(text="处理出错")
                messagebox.showerror("错误", str(e))
            finally:
                self.btn_run.configure(state=tk.NORMAL)

        self.after(100, task)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
